# ============================================================
# APP.PY - File chính của ứng dụng Portal BAU
# Chức năng:
#   1. Xử lý merge file D2C + Result → export CSV + ZIP
#   2. Update lại kết quả khi dev trả file sai
#   3. Xem lịch sử, tìm kiếm, export lại từ history
#   4. Xem log xử lý
# Cách chạy: py app.py → mở http://localhost:5000
# ============================================================

from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import os
import shutil
import zipfile
import pyzipper
import time
from datetime import datetime
from pathlib import Path
from config import (
    BASE_DIR,
    UPLOAD_FOLDER,
    OUTPUT_FOLDER,
    HISTORY_FOLDER,
    HISTORY_DB,
    LOG_FILE,
    ALLOWED_EXTENSIONS,
    NID_VARIANTS,
    RESULT_VARIANTS,
    PHONE_VARIANTS,
    TICKET_VARIANTS,
    USERNAME_VARIANTS,
    PHONE_STANDARD_LENGTH,
    NID_STANDARD_LENGTH,
)
from database import (
    init_db,
    check_ticket_exists,
    save_history_records,
    save_log_records,
    search_history,
    get_log,
    get_all_ticket_names,
    get_records_by_ticket,
    update_it_result_by_ticket,
    update_history_record,
    delete_history_record,
    delete_history_by_ticket,
    get_history_record_by_id,
    delete_history_by_tickets,
    init_responses_db,
    get_all_responses,
    get_all_flows,
    add_response,
    update_response,
    delete_response,
    increment_copy_count,
    get_response_by_id,
    export_all_responses,
    import_responses,
    init_api_catalog_db,
    get_all_api_entries,
    get_all_api_categories,
    add_api_entry,
    get_api_entry_by_id,
    update_api_entry,
    delete_api_entry,
    add_api_image,
    delete_api_image,
    get_images_by_api_id,
    update_api_sort_order,
    export_all_api_entries,
    import_api_entries,
)

# Khởi tạo Flask app
app = Flask(__name__)
# Giới hạn upload tối đa 50MB
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


# Global error handler: luôn trả JSON thay vì trang HTML lỗi
@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    traceback.print_exc()
    return jsonify({"success": False, "error": f"Lỗi server: {str(e)}"}), 500


@app.errorhandler(413)
def handle_file_too_large(e):
    return jsonify({"success": False, "error": "File quá lớn (tối đa 50MB)"}), 413

# Tạo thư mục cần thiết nếu chưa có
for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER, HISTORY_FOLDER]:
    os.makedirs(folder, exist_ok=True)

# Khởi tạo database (tạo bảng + index lần đầu)
init_db()
init_responses_db()
init_api_catalog_db()


# ============================================================
# HÀM KIỂM TRA ĐỊNH DẠNG FILE UPLOAD
# Chỉ cho phép: xlsx, xls, csv
# ============================================================
def allowed_file(filename):
    """Kiểm tra file có đúng định dạng cho phép không."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================
# HÀM TỰ ĐỘNG NHẬN DIỆN TÊN CỘT
# Không cần đúng chính xác tên cột, tool tự hiểu
# VD: "nid", "NID", "CustomerNID" → đều là cột NID
# ============================================================
def detect_column(df, variants):
    """
    Tìm tên cột trong DataFrame khớp với danh sách biến thể.
    Không phân biệt hoa thường, bỏ qua dấu cách và gạch dưới.
    """
    for col in df.columns:
        col_clean = col.strip().lower().replace(" ", "").replace("_", "")
        for v in variants:
            v_clean = v.strip().lower().replace(" ", "").replace("_", "")
            if col_clean == v_clean:
                return col
    return None


# ============================================================
# HÀM CHUẨN HÓA SỐ 0 ĐẦU CHO PHONE VÀ NID
# Phone Việt Nam chuẩn = 10 số (bắt đầu bằng 0)
# NID chuẩn = 12 số (bắt đầu bằng 0)
# Chỉ thêm 0 khi thiếu đúng 1 số, còn lại giữ nguyên
# ============================================================
def fix_leading_zero_phone(value):
    """
    Chuẩn hóa số phone: thêm 0 đầu nếu thiếu so với độ dài chuẩn.
    VD: PHONE_STANDARD_LENGTH = 10
        '961338170' (9 số) → '0961338170' (10 số)
        '0961338170' (10 số) → giữ nguyên
        '12345678' (8 số) → giữ nguyên (lệch quá nhiều)
    Nếu PHONE_STANDARD_LENGTH = 0 → không fix, giữ nguyên.
    """
    if PHONE_STANDARD_LENGTH == 0:
        return value
    if not value or not value.isdigit():
        return value
    # Chỉ thêm 0 khi thiếu đúng 1 số so với chuẩn
    if len(value) == PHONE_STANDARD_LENGTH - 1:
        return "0" + value
    return value


def fix_leading_zero_nid(value):
    """
    Chuẩn hóa NID: thêm 0 đầu nếu thiếu so với độ dài chuẩn.
    VD: NID_STANDARD_LENGTH = 12
        '25200006000' (11 số) → '025200006000' (12 số)
        '025200006000' (12 số) → giữ nguyên
        '123456789' (9 số) → giữ nguyên (lệch quá nhiều)
    Nếu NID_STANDARD_LENGTH = 0 → không fix, giữ nguyên.
    """
    if NID_STANDARD_LENGTH == 0:
        return value
    if not value or not value.isdigit():
        return value
    # Chỉ thêm 0 khi thiếu đúng 1 số so với chuẩn
    if len(value) == NID_STANDARD_LENGTH - 1:
        return "0" + value
    return value


# ============================================================
# HÀM ĐỌC FILE EXCEL/CSV
# Giữ nguyên giá trị gốc - KHÔNG tự cắt số 0 đầu của NID/Phone
# VD: 025200006000 giữ nguyên, không thành 25200006000
# ============================================================
def read_file(filepath):
    """
    Đọc file xlsx/xls/csv thành DataFrame.
    Tất cả giá trị giữ nguyên dạng text gốc từ file.
    """
    ext = filepath.rsplit(".", 1)[1].lower()
    if ext == "csv":
        # CSV: đọc tất cả cột là text, không tự đoán kiểu dữ liệu
        df = pd.read_csv(filepath, dtype=str, keep_default_na=False)
    else:
        # Excel: dùng openpyxl đọc từng cell để giữ nguyên giá trị
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Đọc header (dòng 1)
        headers = []
        for cell in ws[1]:
            headers.append(
                str(cell.value) if cell.value is not None else f"Col_{cell.column}"
            )

        # Đọc data (từ dòng 2 trở đi)
        data = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_data = []
            for cell in row:
                if cell.value is None:
                    row_data.append("")
                else:
                    val = cell.value
                    # Số float kiểu 25200006000.0 → chuyển thành "25200006000"
                    if isinstance(val, float) and val == int(val):
                        row_data.append(str(int(val)))
                    else:
                        row_data.append(str(val))
            data.append(row_data)

        wb.close()
        df = pd.DataFrame(data, columns=headers)

        # Loại bỏ cột trống (không có header thật, chỉ là Col_X)
        real_cols = [c for c in df.columns if not c.startswith('Col_')]
        df = df[real_cols]

    return df


# ============================================================
# HÀM MERGE D2C + RESULT
# Ghép kết quả từ file Result vào file D2C dựa theo NID
# Thêm cột "IT Result" và "Ticket xử lý" rồi export CSV
# ============================================================
def process_merge(d2c_path, result_paths, index, folder_name=None):
    """
    Merge nhiều file Result vào file D2C theo NID.
    result_paths: list các đường dẫn file Result
    Trả về: (dict kết quả, None) nếu thành công
             (None, chuỗi lỗi) nếu thất bại
    """
    # --- Đọc file D2C ---
    try:
        df_d2c = read_file(d2c_path)
    except Exception as e:
        return None, f"File D2C {index}: Không đọc được - {str(e)}"

    # --- Đọc và gộp tất cả file Result ---
    df_results_list = []
    for idx, rpath in enumerate(result_paths, 1):
        try:
            df_r = read_file(rpath)
            df_results_list.append(df_r)
        except Exception as e:
            return None, f"File Result {index} (file {idx}): Không đọc được - {str(e)}"

    # Kiểm tra tất cả file Result phải có cột NID và Result
    for idx, df_r in enumerate(df_results_list, 1):
        nid_col_r = detect_column(df_r, NID_VARIANTS)
        result_col_r = detect_column(df_r, RESULT_VARIANTS)
        if not nid_col_r:
            return None, f"File Result {index} (file {idx}): Thiếu cột NID"
        if not result_col_r:
            return None, f"File Result {index} (file {idx}): Thiếu cột result"

    # --- Tìm cột NID trong D2C ---
    nid_col_d2c = detect_column(df_d2c, NID_VARIANTS)
    if not nid_col_d2c:
        return None, f"File D2C {index}: Thiếu cột NID"

    # --- Loại bỏ dòng trống (dòng không có NID) ---
    df_d2c = df_d2c[df_d2c[nid_col_d2c].str.strip() != ""].reset_index(drop=True)

    # --- Chuẩn hóa số 0 đầu cho NID và Phone trong D2C ---
    df_d2c[nid_col_d2c] = df_d2c[nid_col_d2c].apply(fix_leading_zero_nid)

    # Strip whitespace + loại ký tự ẩn để đảm bảo match
    df_d2c[nid_col_d2c] = df_d2c[nid_col_d2c].str.strip().str.replace(r'\s+', '', regex=True)

    # Chuẩn hóa NID: bỏ số 0 thừa ở đầu để so sánh
    df_d2c['_nid_merge'] = df_d2c[nid_col_d2c].str.lstrip('0')

    # Fix phone trong D2C nếu có cột phone
    phone_col_temp = detect_column(df_d2c, PHONE_VARIANTS)
    if phone_col_temp:
        df_d2c[phone_col_temp] = df_d2c[phone_col_temp].apply(fix_leading_zero_phone)

    # --- Gộp tất cả file Result thành 1 bảng tra cứu NID → Result ---
    # File sau sẽ ghi đè NID trùng từ file trước
    warnings = []
    result_map = {}
    for idx, df_r in enumerate(df_results_list, 1):
        nid_col_r = detect_column(df_r, NID_VARIANTS)
        result_col_r = detect_column(df_r, RESULT_VARIANTS)

        # Loại bỏ dòng trống
        df_r = df_r[df_r[nid_col_r].str.strip() != ""].reset_index(drop=True)
        # Chuẩn hóa NID
        df_r[nid_col_r] = df_r[nid_col_r].apply(fix_leading_zero_nid)
        df_r[nid_col_r] = df_r[nid_col_r].str.strip().str.replace(r'\s+', '', regex=True)
        df_r['_nid_merge'] = df_r[nid_col_r].str.lstrip('0')

        # Kiểm tra NID trùng trong file này
        duplicates = df_r[df_r.duplicated(subset=[nid_col_r], keep=False)]
        if not duplicates.empty:
            dup_nids = duplicates[nid_col_r].unique().tolist()[:5]
            warnings.append(
                f"WARNING: NID trùng trong Result {index} (file {idx}): "
                f"{', '.join(str(x) for x in dup_nids)}"
            )

        # Gộp vào result_map (file sau ghi đè file trước nếu trùng NID)
        file_map = (
            df_r.drop_duplicates(subset=['_nid_merge'], keep="last")
            .set_index('_nid_merge')[result_col_r]
            .to_dict()
        )
        result_map.update(file_map)

    # --- Merge: thêm cột IT Result vào D2C ---
    # NID không tìm thấy trong Result → dừng process, trả lỗi
    df_d2c["IT Result"] = df_d2c['_nid_merge'].map(result_map).fillna("NOT FOUND")

    # Kiểm tra: nếu có NID không tìm thấy → ngưng process
    not_found_df = df_d2c[df_d2c["IT Result"] == "NOT FOUND"]
    if not not_found_df.empty:
        not_found_nids = not_found_df[nid_col_d2c].tolist()
        # Hiển thị tối đa 10 NID đầu tiên
        show_nids = not_found_nids[:10]
        remaining = len(not_found_nids) - 10 if len(not_found_nids) > 10 else 0
        error_msg = (
            f"File D2C {index}: Có {len(not_found_nids)} NID không tìm thấy trong file Result. "
            f"NID: {', '.join(show_nids)}"
        )
        if remaining > 0:
            error_msg += f" ... và {remaining} NID khác"
        # Xóa cột tạm
        df_d2c = df_d2c.drop(columns=['_nid_merge'], errors='ignore')
        return None, error_msg

    # Xóa cột tạm
    df_d2c = df_d2c.drop(columns=['_nid_merge'], errors='ignore')

    # --- Thêm cột "Ticket xử lý" = tên export người dùng đặt ---
    ticket_name = folder_name if folder_name else f"Ticket_{index}"
    df_d2c["Ticket xử lý"] = ticket_name

    # --- Tìm các cột khác để lưu vào history ---
    ticket_col = detect_column(df_d2c, TICKET_VARIANTS)
    phone_col = detect_column(df_d2c, PHONE_VARIANTS)
    username_col = detect_column(df_d2c, USERNAME_VARIANTS)
    d2c_result_col = detect_column(
        df_d2c, RESULT_VARIANTS + ["d2c result", "d2cresult"]
    )

    # --- Sắp xếp cột: "Ticket xử lý" đưa lên đầu ---
    cols = df_d2c.columns.tolist()
    if "Ticket xử lý" in cols:
        cols.remove("Ticket xử lý")
        cols = ["Ticket xử lý"] + cols
        df_d2c = df_d2c[cols]

    # --- Tạo thư mục output và export Excel (.xlsx) ---
    ticket_folder = os.path.join(OUTPUT_FOLDER, ticket_name)
    os.makedirs(ticket_folder, exist_ok=True)
    xlsx_filename = f"{ticket_name}.xlsx"
    xlsx_path = os.path.join(ticket_folder, xlsx_filename)
    # Export Excel với tất cả cột format Text (giữ số 0 đầu, không bị scientific notation)
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df_d2c.to_excel(writer, index=False, sheet_name='Data')
        worksheet = writer.sheets['Data']
        for col in worksheet.columns:
            for cell in col:
                cell.number_format = '@'  # Format text

    # --- Chuẩn bị dữ liệu lưu vào history ---
    history_records = []
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    for _, row in df_d2c.iterrows():
        record = {
            "NewPhoneNumber": row.get(phone_col, "") if phone_col else "",
            "NID": row.get(nid_col_d2c, ""),
            "NewUsername": row.get(username_col, "") if username_col else "",
            "TicketPega": row.get(ticket_col, "") if ticket_col else "",
            "Ticket xử lý": ticket_name,
            "Date Update": now,
            "D2C Result": row.get(d2c_result_col, "") if d2c_result_col else "",
            "IT Result": row.get("IT Result", ""),
        }
        history_records.append(record)

    # --- Trả về kết quả ---
    return {
        "ticket_name": ticket_name,
        "ticket_folder": ticket_folder,
        "csv_path": xlsx_path,
        "history_records": history_records,
        "warnings": warnings,
        "total_rows": len(df_d2c),
        "success_count": len(df_d2c[df_d2c["IT Result"] != "NOT FOUND"]),
        "not_found_count": len(df_d2c[df_d2c["IT Result"] == "NOT FOUND"]),
    }, None


# ============================================================
# HÀM TẠO FILE ZIP
# Có 2 chế độ: có password (AES-256) hoặc không password
# ============================================================
def create_zip(folder_path, zip_path, password=None):
    """
    Nén thư mục thành file ZIP.
    Nếu có password → mã hóa AES-256 (cần pyzipper).
    Nếu không password → ZIP bình thường.
    """
    if password:
        # ZIP có password - mã hóa AES
        with pyzipper.AESZipFile(
            zip_path, "w", compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES
        ) as zf:
            zf.setpassword(password.encode("utf-8"))
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, os.path.dirname(folder_path))
                    zf.write(file_path, arcname)
    else:
        # ZIP không password
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, os.path.dirname(folder_path))
                    zf.write(file_path, arcname)


# ============================================================
# HÀM LƯU LỊCH SỬ VÀ LOG VÀO DATABASE
# ============================================================
def save_history(records):
    """Chuyển đổi format rồi lưu vào SQLite."""
    db_records = []
    for r in records:
        db_records.append(
            {
                "NewPhoneNumber": r.get("NewPhoneNumber", ""),
                "NID": r.get("NID", ""),
                "NewUsername": r.get("NewUsername", ""),
                "TicketPega": r.get("TicketPega", ""),
                "TicketXuLy": r.get("Ticket xử lý", ""),
                "DateUpdate": r.get("Date Update", ""),
                "D2CResult": r.get("D2C Result", ""),
                "ITResult": r.get("IT Result", ""),
            }
        )
    save_history_records(db_records)


def save_log(log_entries):
    """Lưu log xử lý vào database."""
    save_log_records(log_entries)


# ============================================================
# ROUTES - CÁC TRANG GIAO DIỆN
# ============================================================


@app.route("/")
def index():
    """Trang chính: Xử lý Update Phone (gom Process + Update)."""
    return render_template("index.html")


@app.route("/history")
def history_page():
    """Trang lịch sử: tìm kiếm NID/Phone/Ticket + export."""
    return render_template("history.html")


@app.route("/responses")
def responses_page():
    """Trang mẫu phản hồi: hiển thị các câu trả lời mẫu theo luồng."""
    return render_template("responses.html")


# ============================================================
# API - MẪU PHẢN HỒI (CRUD + Export/Import)
# ============================================================

@app.route("/api/responses", methods=["GET"])
def api_get_responses():
    """Lấy danh sách mẫu phản hồi, lọc theo flow và search."""
    flow = request.args.get("flow", "")
    search = request.args.get("search", "")
    data = get_all_responses(flow, search)
    flows = get_all_flows()
    return jsonify({"success": True, "data": data, "flows": flows})


@app.route("/api/responses", methods=["POST"])
def api_add_response():
    """Thêm mẫu phản hồi mới."""
    body = request.get_json()
    title = body.get("title", "").strip()
    flow = body.get("flow", "").strip()
    content = body.get("content", "").strip()
    tags = body.get("tags", "").strip()

    if not title or not flow or not content:
        return jsonify({"success": False, "error": "Thiếu title, flow hoặc content"}), 400

    new_id = add_response(title, flow, content, tags)
    return jsonify({"success": True, "id": new_id, "message": "Đã thêm mẫu phản hồi"})


@app.route("/api/responses/<int:response_id>", methods=["PUT"])
def api_update_response(response_id):
    """Cập nhật mẫu phản hồi."""
    body = request.get_json()
    title = body.get("title", "").strip()
    flow = body.get("flow", "").strip()
    content = body.get("content", "").strip()
    tags = body.get("tags", "").strip()

    if not title or not flow or not content:
        return jsonify({"success": False, "error": "Thiếu title, flow hoặc content"}), 400

    ok = update_response(response_id, title, flow, content, tags)
    if ok:
        return jsonify({"success": True, "message": "Đã cập nhật"})
    return jsonify({"success": False, "error": "Không tìm thấy mẫu"}), 404


@app.route("/api/responses/<int:response_id>", methods=["DELETE"])
def api_delete_response(response_id):
    """Xóa mẫu phản hồi."""
    ok = delete_response(response_id)
    if ok:
        return jsonify({"success": True, "message": "Đã xóa"})
    return jsonify({"success": False, "error": "Không tìm thấy mẫu"}), 404


@app.route("/api/responses/<int:response_id>/copy", methods=["POST"])
def api_copy_response(response_id):
    """Tăng copy count khi user copy mẫu."""
    increment_copy_count(response_id)
    return jsonify({"success": True})


@app.route("/api/responses/export", methods=["GET"])
def api_export_responses():
    """Export toàn bộ mẫu phản hồi ra JSON."""
    import json
    data = export_all_responses()
    # Tạo file JSON tạm
    export_path = os.path.join(OUTPUT_FOLDER, "response_templates_backup.json")
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    with open(export_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # Ghi log
    save_log_records([{
        "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "File": "EXPORT RESPONSES",
        "Status": "SUCCESS",
        "Error": f"Exported {len(data)} mẫu phản hồi ra JSON",
        "Duration": "",
    }])

    return send_file(export_path, as_attachment=True, download_name="response_templates_backup.json")


@app.route("/api/responses/import", methods=["POST"])
def api_import_responses():
    """Import mẫu phản hồi từ file JSON."""
    import json
    if "file" not in request.files:
        return jsonify({"success": False, "error": "Chưa chọn file"}), 400

    file = request.files["file"]
    if not file.filename.endswith(".json"):
        return jsonify({"success": False, "error": "Chỉ hỗ trợ file .json"}), 400

    try:
        content = file.read().decode("utf-8")
        records = json.loads(content)
        imported, skipped = import_responses(records)
        return jsonify({
            "success": True,
            "message": f"Đã import {imported} mẫu phản hồi, bỏ qua {skipped} mẫu trùng"
        })
    except Exception as e:
        return jsonify({"success": False, "error": f"Lỗi đọc file: {str(e)}"}), 400


@app.route("/log")
def log_page():
    """Trang log: xem lịch sử các lần chạy tool."""
    return render_template("log.html")


# ============================================================
# API - EXPORT TOÀN BỘ HISTORY RA CSV
# ============================================================
@app.route("/api/export-all-history")
def export_all_history():
    """Export toàn bộ bảng history ra file Excel (.xlsx) giữ nguyên số 0 đầu."""
    import sqlite3
    from config import HISTORY_DB

    conn = sqlite3.connect(HISTORY_DB)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM history ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        return jsonify({"error": "Không có dữ liệu trong history"}), 404

    data = []
    for row in rows:
        data.append({
            "Ticket xử lý": row["TicketXuLy"],
            "NewPhoneNumber": row["NewPhoneNumber"],
            "NID": row["NID"],
            "NewUsername": row["NewUsername"],
            "TicketPega": row["TicketPega"],
            "Date Update": row["DateUpdate"],
            "D2C Result": row["D2CResult"],
            "IT Result": row["ITResult"],
        })

    df = pd.DataFrame(data)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    export_path = os.path.join(OUTPUT_FOLDER, "All_History_Export.xlsx")

    # Export Excel với tất cả cột là text (giữ số 0 đầu)
    with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='History')
        worksheet = writer.sheets['History']
        for col in worksheet.columns:
            for cell in col:
                cell.number_format = '@'  # Format text

    # Ghi log
    save_log_records([{
        "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "File": "EXPORT ALL HISTORY",
        "Status": "SUCCESS",
        "Error": f"Exported {len(data)} records từ toàn bộ history",
        "Duration": "",
    }])

    return send_file(export_path, as_attachment=True, download_name="All_History_Export.xlsx")


@app.route("/api/export-selected-history")
def export_selected_history():
    """Export các ticket đã chọn ra file Excel (.xlsx) giữ nguyên số 0 đầu."""
    import sqlite3
    from config import HISTORY_DB

    tickets = request.args.getlist("tickets")
    if not tickets:
        return jsonify({"error": "Chưa chọn ticket nào"}), 400

    conn = sqlite3.connect(HISTORY_DB)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    placeholders = ",".join(["?" for _ in tickets])
    cursor.execute(
        f"SELECT * FROM history WHERE TicketXuLy IN ({placeholders}) ORDER BY id DESC",
        tickets
    )
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        return jsonify({"error": "Không tìm thấy dữ liệu cho các ticket đã chọn"}), 404

    data = []
    for row in rows:
        data.append({
            "Ticket xử lý": row["TicketXuLy"],
            "NewPhoneNumber": row["NewPhoneNumber"],
            "NID": row["NID"],
            "NewUsername": row["NewUsername"],
            "TicketPega": row["TicketPega"],
            "Date Update": row["DateUpdate"],
            "D2C Result": row["D2CResult"],
            "IT Result": row["ITResult"],
        })

    df = pd.DataFrame(data)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    export_path = os.path.join(OUTPUT_FOLDER, "Selected_History_Export.xlsx")

    # Export Excel với tất cả cột là text (giữ số 0 đầu)
    with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='History')
        worksheet = writer.sheets['History']
        for col in worksheet.columns:
            for cell in col:
                cell.number_format = '@'  # Format text

    # Ghi log
    save_log_records([{
        "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "File": "EXPORT SELECTED",
        "Status": "SUCCESS",
        "Error": f"Exported {len(data)} records | Tickets: {', '.join(tickets)}",
        "Duration": "",
    }])

    return send_file(export_path, as_attachment=True, download_name="Selected_History_Export.xlsx")


# ============================================================
# API - IMPORT HISTORY TỪ FILE EXCEL CÁ NHÂN
# Cho phép import file Excel cũ vào bảng history
# Tự nhận diện cột dựa theo tên (không cần chính xác)
# ============================================================
@app.route("/api/import-history", methods=["POST"])
def import_history():
    """Import file Excel/CSV vào bảng history."""
    if "file" not in request.files or request.files["file"].filename == "":
        return jsonify({"success": False, "error": "Chưa chọn file"}), 400

    file = request.files["file"]
    if not allowed_file(file.filename):
        return jsonify({"success": False, "error": "Chỉ hỗ trợ xlsx, xls, csv"}), 400

    # Lưu file tạm
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    filepath = os.path.join(UPLOAD_FOLDER, f"import_{file.filename}")
    file.save(filepath)

    try:
        df = read_file(filepath)
    except Exception as e:
        return jsonify({"success": False, "error": f"Không đọc được file: {str(e)}"}), 400

    # Nhận diện các cột trong file
    # Cột NID
    nid_col = detect_column(df, NID_VARIANTS)
    # Cột Phone
    phone_col = detect_column(df, PHONE_VARIANTS)
    # Cột Username
    username_col = detect_column(df, USERNAME_VARIANTS)
    # Cột TicketPega
    ticket_col = detect_column(df, TICKET_VARIANTS)
    # Cột Ticket xử lý (ID Ticket, TicketXuLy, Ticket xử lý)
    ticket_xuly_col = detect_column(df, ['id ticket', 'idticket', 'ticket xử lý', 'ticketxuly', 'ticket xu ly'])
    # Cột Date
    date_col = detect_column(df, ['update phone(day)', 'updatephone', 'date update', 'dateupdate', 'date', 'ngày'])
    # Cột D2C Result
    d2c_col = detect_column(df, ['d2c result', 'd2cresult', 'd2c_result'])
    # Cột IT Result/Respond
    it_col = detect_column(df, ['it respond', 'itrespond', 'it result', 'itresult', 'it_result', 'it_respond'])

    if not nid_col and not phone_col:
        return jsonify({"success": False, "error": "File phải có ít nhất cột NID hoặc Phone"}), 400

    # Loại bỏ dòng trống
    if nid_col:
        df = df[df[nid_col].str.strip() != ""].reset_index(drop=True)
    elif phone_col:
        df = df[df[phone_col].str.strip() != ""].reset_index(drop=True)

    # Chuẩn hóa số 0 đầu
    if nid_col:
        df[nid_col] = df[nid_col].apply(fix_leading_zero_nid)
    if phone_col:
        df[phone_col] = df[phone_col].apply(fix_leading_zero_phone)

    # Tạo records để lưu vào history
    records = []
    for _, row in df.iterrows():
        record = {
            "NewPhoneNumber": row.get(phone_col, "") if phone_col else "",
            "NID": row.get(nid_col, "") if nid_col else "",
            "NewUsername": row.get(username_col, "") if username_col else "",
            "TicketPega": row.get(ticket_col, "") if ticket_col else "",
            "TicketXuLy": row.get(ticket_xuly_col, "") if ticket_xuly_col else "",
            "DateUpdate": row.get(date_col, "") if date_col else "",
            "D2CResult": row.get(d2c_col, "") if d2c_col else "",
            "ITResult": row.get(it_col, "") if it_col else "",
        }
        records.append(record)

    # Lưu vào database
    if records:
        save_history_records(records)

    # Ghi log
    save_log_records([{
        "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "File": f"IMPORT - {file.filename}",
        "Status": "SUCCESS",
        "Error": f"Imported {len(records)} records",
        "Duration": "",
    }])

    # Cleanup
    try:
        os.remove(filepath)
    except:
        pass

    return jsonify({
        "success": True,
        "message": f"Đã import {len(records)} records vào history",
        "details": {
            "total_rows": len(records),
            "columns_detected": {
                "NID": nid_col or "Không tìm thấy",
                "Phone": phone_col or "Không tìm thấy",
                "Username": username_col or "Không tìm thấy",
                "TicketPega": ticket_col or "Không tìm thấy",
                "Ticket xử lý": ticket_xuly_col or "Không tìm thấy",
                "Date": date_col or "Không tìm thấy",
                "D2C Result": d2c_col or "Không tìm thấy",
                "IT Result": it_col or "Không tìm thấy",
            }
        }
    })


# ============================================================
# API - XỬ LÝ MERGE (Process mới)
# Nhận file D2C + Result → merge theo NID → export CSV + ZIP
# Kiểm tra: tên export bắt buộc, không trùng trong history
# ============================================================
@app.route("/api/process", methods=["POST"])
def process_files():
    """API xử lý merge nhiều cặp file D2C + Result."""
    start_time = time.time()

    # Xóa output cũ, tạo mới
    if os.path.exists(OUTPUT_FOLDER):
        shutil.rmtree(OUTPUT_FOLDER)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Xóa upload cũ
    if os.path.exists(UPLOAD_FOLDER):
        shutil.rmtree(UPLOAD_FOLDER)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

    # Lấy thông tin từ form
    password = request.form.get("password", "").strip()
    file_count = int(request.form.get("file_count", 0))

    if file_count == 0:
        return jsonify({"success": False, "error": "Chưa chọn số lượng file"}), 400

    # Biến lưu kết quả
    results = []
    all_history = []
    all_warnings = []
    success_count = 0
    fail_count = 0
    log_entries = []
    zip_files = []

    # --- Xử lý từng cặp file ---
    for i in range(1, file_count + 1):
        d2c_key = f"d2c_{i}"
        result_key = f"result_{i}"
        export_name = request.form.get(f"export_name_{i}", "").strip()

        # Kiểm tra: phải có tên export
        if not export_name:
            fail_count += 1
            results.append(
                {
                    "index": i,
                    "status": "error",
                    "message": f"Chưa đặt tên file export số {i}",
                }
            )
            log_entries.append(
                {
                    "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "File": f"Pair {i}",
                    "Status": "FAIL",
                    "Error": "Thiếu tên export",
                    "Duration": "",
                }
            )
            continue

        # Kiểm tra: tên export đã có trong history → không cho chạy trùng
        if check_ticket_exists(export_name):
            fail_count += 1
            results.append(
                {
                    "index": i,
                    "status": "error",
                    "message": f'Tên "{export_name}" đã tồn tại trong lịch sử!',
                }
            )
            log_entries.append(
                {
                    "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "File": f"Pair {i}",
                    "Status": "FAIL",
                    "Error": f"Trùng: {export_name}",
                    "Duration": "",
                }
            )
            continue

        # Kiểm tra: file D2C có được upload không
        if d2c_key not in request.files or request.files[d2c_key].filename == "":
            fail_count += 1
            results.append(
                {"index": i, "status": "error", "message": f"Thiếu file D2C số {i}"}
            )
            log_entries.append(
                {
                    "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "File": f"D2C {i}",
                    "Status": "FAIL",
                    "Error": "Thiếu file D2C",
                    "Duration": "",
                }
            )
            continue

        # Kiểm tra: file Result có được upload không (hỗ trợ nhiều file)
        result_files = request.files.getlist(result_key)
        if not result_files or all(f.filename == "" for f in result_files):
            fail_count += 1
            results.append(
                {"index": i, "status": "error", "message": f"Thiếu file Result số {i}"}
            )
            log_entries.append(
                {
                    "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "File": f"Result {i}",
                    "Status": "FAIL",
                    "Error": "Thiếu file Result",
                    "Duration": "",
                }
            )
            continue

        d2c_file = request.files[d2c_key]

        # Kiểm tra định dạng file D2C
        if not allowed_file(d2c_file.filename):
            fail_count += 1
            results.append(
                {
                    "index": i,
                    "status": "error",
                    "message": f"File D2C {i}: Chỉ hỗ trợ xlsx, xls, csv",
                }
            )
            continue

        # Kiểm tra định dạng tất cả file Result
        invalid_result = False
        for rf in result_files:
            if rf.filename and not allowed_file(rf.filename):
                fail_count += 1
                results.append(
                    {
                        "index": i,
                        "status": "error",
                        "message": f"File Result {i} ({rf.filename}): Chỉ hỗ trợ xlsx, xls, csv",
                    }
                )
                invalid_result = True
                break
        if invalid_result:
            continue

        # Lưu file upload tạm
        d2c_path = os.path.join(UPLOAD_FOLDER, f"d2c_{i}_{d2c_file.filename}")
        d2c_file.save(d2c_path)

        # Lưu tất cả file Result
        result_paths = []
        for j, rf in enumerate(result_files):
            if rf.filename:
                rpath = os.path.join(UPLOAD_FOLDER, f"result_{i}_{j}_{rf.filename}")
                rf.save(rpath)
                result_paths.append(rpath)

        # --- Thực hiện merge (truyền list các file Result) ---
        merge_start = time.time()
        folder_name = export_name if export_name else f"Ticket_{i}"
        merge_result, error = process_merge(d2c_path, result_paths, i, folder_name)
        merge_duration = round(time.time() - merge_start, 2)

        if error:
            fail_count += 1
            results.append({"index": i, "status": "error", "message": error})
            log_entries.append(
                {
                    "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "File": f"Pair {i}",
                    "Status": "FAIL",
                    "Error": error,
                    "Duration": f"{merge_duration}s",
                }
            )
            continue

        # --- Tạo file ZIP ---
        zip_filename = f"{merge_result['ticket_name']}.zip"
        zip_path = os.path.join(OUTPUT_FOLDER, zip_filename)
        create_zip(
            merge_result["ticket_folder"], zip_path, password if password else None
        )
        zip_files.append(zip_filename)

        # Gom lịch sử + cảnh báo
        all_history.extend(merge_result["history_records"])
        all_warnings.extend(merge_result["warnings"])

        # Ghi nhận thành công
        success_count += 1
        results.append(
            {
                "index": i,
                "status": "success",
                "message": (
                    f'OK - {merge_result["total_rows"]} rows, '
                    f'{merge_result["success_count"]} matched, '
                    f'{merge_result["not_found_count"]} not found'
                ),
                "zip_file": zip_filename,
                "warnings": merge_result["warnings"],
            }
        )
        log_entries.append(
            {
                "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "File": f"PROCESS - Ticket: {folder_name} | D2C: {d2c_file.filename} | Result: {', '.join(rf.filename for rf in result_files if rf.filename)}",
                "Status": "SUCCESS",
                "Error": f'{merge_result["total_rows"]} rows, {merge_result["success_count"]} matched, {merge_result["not_found_count"]} not found',
                "Duration": f"{merge_duration}s",
            }
        )

    # --- Lưu history và log ---
    if all_history:
        save_history(all_history)
    if log_entries:
        save_log(log_entries)

    total_time = round(time.time() - start_time, 2)

    # Trả về kết quả cho frontend
    return jsonify(
        {
            "success": True,
            "results": results,
            "summary": {
                "total": file_count,
                "success": success_count,
                "fail": fail_count,
                "time": f"{total_time}s",
                "password": password if password else "Không",
            },
            "warnings": all_warnings,
            "zip_files": zip_files,
        }
    )


# ============================================================
# API - UPDATE RESULT (Cập nhật kết quả khi dev trả sai)
# Cho phép update nhiều ticket cùng lúc, bao nhiêu lần cũng được
# Upload file Result mới (có NID + result) → update trong history
# ============================================================
@app.route("/api/update-result", methods=["POST"])
def update_result():
    """API cập nhật IT Result cho ticket đã xử lý trước đó."""
    start_time = time.time()

    update_count = int(request.form.get("update_count", 0))
    password = request.form.get("password", "").strip()

    if update_count == 0:
        return jsonify({"success": False, "error": "Chưa có file nào để update"}), 400

    # Thư mục output cho update
    update_output = os.path.join(OUTPUT_FOLDER, "_updates")
    if os.path.exists(update_output):
        shutil.rmtree(update_output)
    os.makedirs(update_output, exist_ok=True)

    results = []
    zip_files = []
    total_updated = 0
    total_not_found = 0

    # --- Xử lý từng ticket cần update ---
    for i in range(1, update_count + 1):
        ticket_name = request.form.get(f"ticket_name_{i}", "").strip()
        file_key = f"result_file_{i}"

        # Kiểm tra chọn ticket chưa
        if not ticket_name:
            results.append(
                {"index": i, "status": "error", "message": f"Chưa chọn ticket số {i}"}
            )
            continue

        # Kiểm tra có file Result không
        if file_key not in request.files or request.files[file_key].filename == "":
            results.append(
                {
                    "index": i,
                    "status": "error",
                    "message": f'Thiếu file Result cho "{ticket_name}"',
                }
            )
            continue

        result_file = request.files[file_key]
        if not allowed_file(result_file.filename):
            results.append(
                {
                    "index": i,
                    "status": "error",
                    "message": f"File {i}: Chỉ hỗ trợ xlsx, xls, csv",
                }
            )
            continue

        # Ticket phải tồn tại trong history mới update được
        if not check_ticket_exists(ticket_name):
            results.append(
                {
                    "index": i,
                    "status": "error",
                    "message": f'Ticket "{ticket_name}" không có trong lịch sử',
                }
            )
            continue

        # Lưu và đọc file
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        result_path = os.path.join(UPLOAD_FOLDER, f"update_{i}_{result_file.filename}")
        result_file.save(result_path)

        try:
            df_result = read_file(result_path)
        except Exception as e:
            results.append(
                {
                    "index": i,
                    "status": "error",
                    "message": f"Không đọc được file: {str(e)}",
                }
            )
            continue

        # Tìm cột NID và Result trong file mới
        nid_col = detect_column(df_result, NID_VARIANTS)
        result_col = detect_column(df_result, RESULT_VARIANTS)
        if not nid_col:
            results.append(
                {"index": i, "status": "error", "message": f"File {i}: Thiếu cột NID"}
            )
            continue
        if not result_col:
            results.append(
                {
                    "index": i,
                    "status": "error",
                    "message": f"File {i}: Thiếu cột result",
                }
            )
            continue

        # Tạo bảng tra cứu NID → Result mới
        nid_result_map = (
            df_result.drop_duplicates(subset=[nid_col], keep="last")
            .set_index(nid_col)[result_col]
            .to_dict()
        )

        # Cập nhật trong database
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        updated, not_found = update_it_result_by_ticket(
            ticket_name, nid_result_map, now
        )
        total_updated += updated
        total_not_found += not_found

        # Export CSV mới sau khi update
        records = get_records_by_ticket(ticket_name)
        if records:
            ticket_folder = os.path.join(update_output, ticket_name)
            os.makedirs(ticket_folder, exist_ok=True)

            df_export = pd.DataFrame(records)
            df_export = df_export.drop(columns=["id"], errors="ignore")
            df_export.columns = [
                "NewPhoneNumber",
                "NID",
                "NewUsername",
                "TicketPega",
                "Ticket xử lý",
                "Date Update",
                "D2C Result",
                "IT Result",
            ]
            # Đưa "Ticket xử lý" lên đầu
            df_export = df_export[
                [
                    "Ticket xử lý",
                    "NewPhoneNumber",
                    "NID",
                    "NewUsername",
                    "TicketPega",
                    "Date Update",
                    "D2C Result",
                    "IT Result",
                ]
            ]
            csv_path = os.path.join(ticket_folder, f"{ticket_name}.xlsx")
            with pd.ExcelWriter(csv_path, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name='Data')
                worksheet = writer.sheets['Data']
                for col in worksheet.columns:
                    for cell in col:
                        cell.number_format = '@'

            # Tạo ZIP
            zip_filename = f"{ticket_name}.zip"
            zip_path = os.path.join(update_output, zip_filename)
            create_zip(ticket_folder, zip_path, password if password else None)
            zip_files.append(zip_filename)

        results.append(
            {
                "index": i,
                "status": "success",
                "message": f'OK - "{ticket_name}": {updated} updated, {not_found} not found',
                "zip_file": f"_updates/{zip_filename}" if records else None,
            }
        )

    total_time = round(time.time() - start_time, 2)

    # Ghi log chi tiết: ticket nào update, file nào, kết quả ra sao
    updated_tickets = [r["message"] for r in results if r["status"] == "success"]
    save_log_records(
        [
            {
                "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "File": f"UPDATE RESULT - {update_count} ticket(s)",
                "Status": "SUCCESS",
                "Error": (
                    f"Tổng: {total_updated} updated, {total_not_found} not found | Chi tiết: {'; '.join(updated_tickets)}"
                    if updated_tickets
                    else "Không có ticket nào update thành công"
                ),
                "Duration": f"{total_time}s",
            }
        ]
    )

    return jsonify(
        {
            "success": True,
            "results": results,
            "summary": {
                "total_updated": total_updated,
                "total_not_found": total_not_found,
                "time": f"{total_time}s",
            },
            "zip_files": zip_files,
        }
    )


# ============================================================
# API - DOWNLOAD FILE
# ============================================================
@app.route("/api/download/<path:filename>")
def download_file(filename):
    """Download file ZIP từ thư mục output."""
    file_path = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return jsonify({"error": "File not found"}), 404


@app.route("/api/download-all")
def download_all():
    """Gom tất cả file ZIP thành 1 file ZIP tổng để download."""
    all_zip_path = os.path.join(OUTPUT_FOLDER, "All_Results.zip")
    with zipfile.ZipFile(all_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in os.listdir(OUTPUT_FOLDER):
            if f.endswith(".zip") and f != "All_Results.zip":
                zf.write(os.path.join(OUTPUT_FOLDER, f), f)
    return send_file(all_zip_path, as_attachment=True)


# ============================================================
# API - EXPORT TICKET TỪ HISTORY
# Tìm ticket trong history → export ra CSV + ZIP
# ============================================================
@app.route("/api/export-ticket/<ticket_name>")
def export_ticket(ticket_name):
    """Export 1 ticket từ history thành file Excel ZIP."""
    records = get_records_by_ticket(ticket_name)
    if not records:
        return jsonify({"error": "Ticket không tồn tại"}), 404

    # Tạo thư mục tạm
    export_folder = os.path.join(OUTPUT_FOLDER, "_export", ticket_name)
    os.makedirs(export_folder, exist_ok=True)

    # Export Excel - "Ticket xử lý" đưa lên đầu
    df_export = pd.DataFrame(records)
    df_export = df_export.drop(columns=["id"], errors="ignore")
    df_export.columns = [
        "NewPhoneNumber",
        "NID",
        "NewUsername",
        "TicketPega",
        "Ticket xử lý",
        "Date Update",
        "D2C Result",
        "IT Result",
    ]
    df_export = df_export[
        [
            "Ticket xử lý",
            "NewPhoneNumber",
            "NID",
            "NewUsername",
            "TicketPega",
            "Date Update",
            "D2C Result",
            "IT Result",
        ]
    ]
    xlsx_path = os.path.join(export_folder, f"{ticket_name}.xlsx")
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='History')
        worksheet = writer.sheets['History']
        for col in worksheet.columns:
            for cell in col:
                cell.number_format = '@'

    # ZIP không password
    zip_path = os.path.join(OUTPUT_FOLDER, "_export", f"{ticket_name}.zip")
    create_zip(export_folder, zip_path, None)

    # Ghi log
    save_log_records([{
        "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "File": "EXPORT TICKET",
        "Status": "SUCCESS",
        "Error": f"Exported {len(records)} records | Ticket: {ticket_name}",
        "Duration": "",
    }])

    return send_file(zip_path, as_attachment=True)


# ============================================================
# API - HISTORY (Lịch sử + tìm kiếm)
# ============================================================
@app.route("/api/history")
def get_history_api():
    """Lấy dữ liệu lịch sử, có hỗ trợ tìm kiếm và phân trang."""
    search_query = request.args.get("search", "").strip()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    data, total = search_history(search_query, page, per_page)
    return jsonify({"data": data, "total": total, "page": page, "per_page": per_page})


# ============================================================
# API - LOG (Lịch sử chạy tool)
# ============================================================
@app.route("/api/log")
def get_log_api():
    """Lấy danh sách log xử lý, có tìm kiếm và phân trang."""
    search_query = request.args.get("search", "").strip()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    data, total = get_log(search_query, page, per_page)
    return jsonify({"data": data, "total": total, "page": page, "per_page": per_page})


@app.route("/api/log/delete/<int:log_id>", methods=["DELETE"])
def delete_log_record(log_id):
    """Xóa 1 dòng log theo ID."""
    from database import get_db

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM process_log WHERE id = ?", (log_id,))
    conn.commit()
    affected = cursor.rowcount
    conn.close()
    if affected > 0:
        return jsonify({"success": True, "message": "Đã xóa log"})
    return jsonify({"success": False, "error": "Không tìm thấy"}), 404


@app.route("/api/log/delete-all", methods=["DELETE"])
def delete_all_logs():
    """Xóa toàn bộ nhật ký."""
    from database import get_db

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM process_log")
    count = cursor.rowcount
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"Đã xóa toàn bộ {count} dòng log"})


@app.route("/api/log/delete-by-search", methods=["DELETE"])
def delete_logs_by_search():
    """Xóa các dòng log khớp với từ khóa tìm kiếm."""
    search = request.args.get("search", "").strip()
    if not search:
        return jsonify({"success": False, "error": "Chưa nhập từ khóa tìm kiếm"}), 400

    from database import get_db

    conn = get_db()
    cursor = conn.cursor()
    like = f"%{search}%"
    cursor.execute(
        """
        DELETE FROM process_log
        WHERE DateTime LIKE ? OR File LIKE ? OR Status LIKE ? OR Error LIKE ?
    """,
        (like, like, like, like),
    )
    count = cursor.rowcount
    conn.commit()
    conn.close()
    return jsonify(
        {"success": True, "message": f'Đã xóa {count} dòng log khớp "{search}"'}
    )


# ============================================================
# API - DANH SÁCH TICKET (dùng cho dropdown Update)
# ============================================================
@app.route("/api/tickets")
def get_tickets():
    """Lấy tất cả tên ticket đã xử lý (cho dropdown)."""
    tickets = get_all_ticket_names()
    return jsonify({"tickets": tickets})


@app.route("/api/ticket-detail/<ticket_name>")
def get_ticket_detail(ticket_name):
    """Lấy chi tiết tất cả records của 1 ticket (preview)."""
    records = get_records_by_ticket(ticket_name)
    return jsonify({"data": records, "total": len(records)})


# ============================================================
# API - CRUD HISTORY (Sửa/Xóa record bằng tay)
# Dùng khi cần chỉnh sửa thông tin sai trong history
# ============================================================


@app.route("/api/history/record/<int:record_id>")
def get_record(record_id):
    """Lấy 1 record theo ID (để hiển thị form sửa)."""
    record = get_history_record_by_id(record_id)
    if record:
        return jsonify({"success": True, "data": record})
    return jsonify({"success": False, "error": "Không tìm thấy record"}), 404


@app.route("/api/history/update/<int:record_id>", methods=["POST"])
def update_record(record_id):
    """
    Cập nhật 1 record trong history (sửa tay).
    Nhận JSON body với các field cần sửa.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "Không có dữ liệu"}), 400

    # Lấy thông tin record TRƯỚC khi sửa (để so sánh cũ → mới)
    old_record = get_history_record_by_id(record_id)
    if not old_record:
        return jsonify({"success": False, "error": "Không tìm thấy record"}), 404

    edit_start = time.time()
    success = update_history_record(record_id, data)
    edit_duration = round(time.time() - edit_start, 2)
    if success:
        # Ghi nhật ký chi tiết: field nào thay đổi từ gì → thành gì
        nid = old_record["NID"]
        ticket = old_record["TicketXuLy"]

        # So sánh giá trị cũ và mới cho từng field
        # Map tên field trong DB sang tên hiển thị
        field_names = {
            "NID": "NID",
            "NewPhoneNumber": "Phone",
            "NewUsername": "Username",
            "TicketPega": "TicketPega",
            "D2CResult": "D2C Result",
            "ITResult": "IT Result",
        }
        changes = []
        for key, new_val in data.items():
            old_val = old_record.get(key, "")
            if old_val != new_val:
                display_name = field_names.get(key, key)
                changes.append(f'{display_name}: "{old_val}" → "{new_val}"')

        change_detail = " | ".join(changes) if changes else "Không có thay đổi"

        save_log_records(
            [
                {
                    "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "File": f"SỬA TAY - Ticket: {ticket} | NID: {nid}",
                    "Status": "SUCCESS",
                    "Error": change_detail,
                    "Duration": f"{edit_duration}s",
                }
            ]
        )
        return jsonify({"success": True, "message": "Đã cập nhật thành công"})
    return jsonify({"success": False, "error": "Cập nhật thất bại"}), 500


@app.route("/api/history/delete/<int:record_id>", methods=["DELETE"])
def delete_record(record_id):
    """Xóa 1 record khỏi history."""
    # Lấy thông tin trước khi xóa
    old_record = get_history_record_by_id(record_id)
    success = delete_history_record(record_id)
    if success:
        nid = old_record["NID"] if old_record else "?"
        ticket = old_record["TicketXuLy"] if old_record else "?"
        phone = old_record["NewPhoneNumber"] if old_record else "?"
        save_log_records(
            [
                {
                    "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "File": f"XÓA RECORD - Ticket: {ticket} | NID: {nid} | Phone: {phone}",
                    "Status": "SUCCESS",
                    "Error": "",
                    "Duration": "",
                }
            ]
        )
        return jsonify({"success": True, "message": "Đã xóa"})
    return jsonify({"success": False, "error": "Không tìm thấy record"}), 404


@app.route("/api/history/delete-ticket/<ticket_name>", methods=["DELETE"])
def delete_ticket(ticket_name):
    """Xóa toàn bộ records của 1 ticket."""
    count = delete_history_by_ticket(ticket_name)
    save_log_records(
        [
            {
                "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "File": f"XÓA TICKET: {ticket_name}",
                "Status": "SUCCESS",
                "Error": f"Đã xóa toàn bộ {count} records thuộc ticket này",
                "Duration": "",
            }
        ]
    )
    return jsonify(
        {
            "success": True,
            "message": f'Đã xóa {count} records của ticket "{ticket_name}"',
        }
    )


@app.route("/api/history/delete-tickets", methods=["POST"])
def delete_tickets():
    """Xóa nhiều ticket cùng lúc."""
    data = request.get_json()
    if not data or "tickets" not in data:
        return jsonify({"success": False, "error": "Thiếu danh sách ticket"}), 400

    ticket_list = data["tickets"]
    if not ticket_list:
        return jsonify({"success": False, "error": "Danh sách rỗng"}), 400

    count = delete_history_by_tickets(ticket_list)
    save_log_records(
        [
            {
                "DateTime": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "File": f'XÓA HÀNG LOẠT: {", ".join(ticket_list)}',
                "Status": "SUCCESS",
                "Error": f"Đã xóa tổng cộng {count} records từ {len(ticket_list)} tickets",
                "Duration": "",
            }
        ]
    )
    return jsonify(
        {
            "success": True,
            "message": f"Đã xóa {count} records từ {len(ticket_list)} tickets",
        }
    )


# ============================================================
# TRANG GIAO DIỆN API CATALOG (Quản lý API kèm hình ảnh)
# ============================================================

@app.route("/image-api")
def image_api_page():
    """Trang quản lý API catalog kèm hình ảnh giao diện."""
    return render_template("imageAPI.html")


# ============================================================
# API - CRUD API CATALOG
# ============================================================

# Thư mục lưu ảnh upload cho API catalog
API_IMAGES_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads', 'api_images')
os.makedirs(API_IMAGES_FOLDER, exist_ok=True)

ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}


def allowed_image(filename):
    """Kiểm tra file ảnh có đúng định dạng không."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


@app.route("/api/image-api", methods=["GET"])
def api_get_api_entries():
    """Lấy danh sách API entries, lọc theo category và search."""
    category = request.args.get("category", "")
    search = request.args.get("search", "")
    data = get_all_api_entries(category, search)
    categories = get_all_api_categories()
    return jsonify({"success": True, "data": data, "categories": categories})


@app.route("/api/image-api", methods=["POST"])
def api_add_api_entry():
    """Thêm API entry mới (multipart/form-data vì có ảnh)."""
    name = request.form.get("name", "").strip()
    method = request.form.get("method", "GET").strip().upper()
    endpoint = request.form.get("endpoint", "").strip()
    description = request.form.get("description", "").strip()
    request_body = request.form.get("request_body", "").strip()
    response_body = request.form.get("response_body", "").strip()
    category = request.form.get("category", "").strip()
    notes = request.form.get("notes", "").strip()
    sort_order = int(request.form.get("sort_order", 0))

    if not name or not endpoint:
        return jsonify({"success": False, "error": "Thiếu tên API hoặc endpoint"}), 400

    # Xử lý upload ảnh chính (thumbnail)
    image_path = ""
    if "image" in request.files:
        image_file = request.files["image"]
        if image_file.filename and allowed_image(image_file.filename):
            import uuid
            ext = image_file.filename.rsplit('.', 1)[1].lower()
            filename = f"{uuid.uuid4().hex}.{ext}"
            image_file.save(os.path.join(API_IMAGES_FOLDER, filename))
            image_path = f"uploads/api_images/{filename}"

    new_id = add_api_entry(name, method, endpoint, description, request_body, response_body, image_path, category, notes, sort_order)

    # Xử lý upload nhiều ảnh phụ
    extra_images = request.files.getlist("extra_images")
    for img_file in extra_images:
        if img_file.filename and allowed_image(img_file.filename):
            import uuid
            ext = img_file.filename.rsplit('.', 1)[1].lower()
            fname = f"{uuid.uuid4().hex}.{ext}"
            img_file.save(os.path.join(API_IMAGES_FOLDER, fname))
            add_api_image(new_id, f"uploads/api_images/{fname}")

    return jsonify({"success": True, "id": new_id, "message": "Đã thêm API entry"})


@app.route("/api/image-api/<int:entry_id>", methods=["PUT"])
def api_update_api_entry(entry_id):
    """Cập nhật API entry (multipart/form-data vì có thể upload ảnh mới)."""
    name = request.form.get("name", "").strip()
    method = request.form.get("method", "GET").strip().upper()
    endpoint = request.form.get("endpoint", "").strip()
    description = request.form.get("description", "").strip()
    request_body = request.form.get("request_body", "").strip()
    response_body = request.form.get("response_body", "").strip()
    category = request.form.get("category", "").strip()
    notes = request.form.get("notes", "").strip()
    sort_order = int(request.form.get("sort_order", 0))

    if not name or not endpoint:
        return jsonify({"success": False, "error": "Thiếu tên API hoặc endpoint"}), 400

    # Xử lý upload ảnh mới (nếu có)
    image_path = None  # None = không thay đổi ảnh
    if "image" in request.files:
        image_file = request.files["image"]
        if image_file.filename and allowed_image(image_file.filename):
            import uuid
            # Xóa ảnh cũ nếu có
            old_entry = get_api_entry_by_id(entry_id)
            if old_entry and old_entry["image_path"]:
                old_image_full = os.path.join(BASE_DIR, 'static', old_entry["image_path"])
                if os.path.exists(old_image_full):
                    os.remove(old_image_full)

            ext = image_file.filename.rsplit('.', 1)[1].lower()
            filename = f"{uuid.uuid4().hex}.{ext}"
            image_file.save(os.path.join(API_IMAGES_FOLDER, filename))
            image_path = f"uploads/api_images/{filename}"

    # Kiểm tra nếu user muốn xóa ảnh (gửi flag remove_image=1)
    if request.form.get("remove_image") == "1":
        old_entry = get_api_entry_by_id(entry_id)
        if old_entry and old_entry["image_path"]:
            old_image_full = os.path.join(BASE_DIR, 'static', old_entry["image_path"])
            if os.path.exists(old_image_full):
                os.remove(old_image_full)
        image_path = ""

    ok = update_api_entry(entry_id, name, method, endpoint, description, request_body, response_body, image_path, category, notes, sort_order)

    # Xử lý upload ảnh phụ mới
    extra_images = request.files.getlist("extra_images")
    for img_file in extra_images:
        if img_file.filename and allowed_image(img_file.filename):
            import uuid
            ext = img_file.filename.rsplit('.', 1)[1].lower()
            fname = f"{uuid.uuid4().hex}.{ext}"
            img_file.save(os.path.join(API_IMAGES_FOLDER, fname))
            add_api_image(entry_id, f"uploads/api_images/{fname}")

    if ok:
        return jsonify({"success": True, "message": "Đã cập nhật"})
    return jsonify({"success": False, "error": "Không tìm thấy API entry"}), 404


@app.route("/api/image-api/<int:entry_id>", methods=["DELETE"])
def api_delete_api_entry(entry_id):
    """Xóa API entry + xóa file ảnh."""
    # Xóa file ảnh chính
    entry = get_api_entry_by_id(entry_id)
    if entry:
        if entry["image_path"]:
            image_full = os.path.join(BASE_DIR, 'static', entry["image_path"])
            if os.path.exists(image_full):
                os.remove(image_full)
        # Xóa file ảnh phụ
        for img in entry.get("images", []):
            img_full = os.path.join(BASE_DIR, 'static', img["image_path"])
            if os.path.exists(img_full):
                os.remove(img_full)

    ok = delete_api_entry(entry_id)
    if ok:
        return jsonify({"success": True, "message": "Đã xóa"})
    return jsonify({"success": False, "error": "Không tìm thấy API entry"}), 404


@app.route("/api/image-api/image/<int:image_id>", methods=["DELETE"])
def api_delete_single_image(image_id):
    """Xóa 1 ảnh phụ."""
    path = delete_api_image(image_id)
    if path:
        img_full = os.path.join(BASE_DIR, 'static', path)
        if os.path.exists(img_full):
            os.remove(img_full)
        return jsonify({"success": True, "message": "Đã xóa ảnh"})
    return jsonify({"success": False, "error": "Không tìm thấy ảnh"}), 404


@app.route("/api/image-api/<int:entry_id>/reorder-images", methods=["POST"])
def api_reorder_images(entry_id):
    """Cập nhật thứ tự ảnh phụ cho 1 API entry."""
    data = request.get_json()
    if not data or "image_ids" not in data:
        return jsonify({"success": False, "error": "Thiếu dữ liệu"}), 400

    from database import get_db
    conn = get_db()
    cursor = conn.cursor()
    for idx, img_id in enumerate(data["image_ids"]):
        cursor.execute('UPDATE api_images SET sort_order = ? WHERE id = ? AND api_id = ?', (idx, img_id, entry_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Đã cập nhật thứ tự ảnh"})


@app.route("/api/image-api/sort", methods=["POST"])
def api_update_sort_order():
    """Cập nhật thứ tự hiển thị cho nhiều API cùng lúc."""
    data = request.get_json()
    if not data or "items" not in data:
        return jsonify({"success": False, "error": "Thiếu dữ liệu"}), 400

    for item in data["items"]:
        update_api_sort_order(item["id"], item["sort_order"])

    return jsonify({"success": True, "message": "Đã cập nhật thứ tự"})


@app.route("/api/image-api/export", methods=["GET"])
def api_export_api_entries():
    """Export toàn bộ API catalog ra ZIP (JSON + ảnh)."""
    import json
    import zipfile as zf

    data = export_all_api_entries()

    # Gom thêm image_path cho từng entry
    all_entries = get_all_api_entries()
    export_data = []
    image_files = []  # (path_in_zip, path_on_disk)

    for entry in all_entries:
        item = {
            'name': entry['name'],
            'method': entry['method'],
            'endpoint': entry['endpoint'],
            'description': entry['description'],
            'request_body': entry['request_body'],
            'response_body': entry['response_body'],
            'category': entry['category'],
            'notes': entry.get('notes', ''),
            'sort_order': entry.get('sort_order', 0),
            'image_path': '',
            'extra_images': []
        }

        # Ảnh chính
        if entry['image_path']:
            full_path = os.path.join(BASE_DIR, 'static', entry['image_path'])
            if os.path.exists(full_path):
                fname = os.path.basename(entry['image_path'])
                item['image_path'] = f"images/{fname}"
                image_files.append((f"images/{fname}", full_path))

        # Ảnh phụ
        for img in entry.get('images', []):
            full_path = os.path.join(BASE_DIR, 'static', img['image_path'])
            if os.path.exists(full_path):
                fname = os.path.basename(img['image_path'])
                item['extra_images'].append(f"images/{fname}")
                image_files.append((f"images/{fname}", full_path))

        export_data.append(item)

    # Tạo ZIP
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    zip_path = os.path.join(OUTPUT_FOLDER, "api_catalog_backup.zip")

    with zf.ZipFile(zip_path, 'w', zf.ZIP_DEFLATED) as zipf:
        # Ghi JSON
        json_content = json.dumps(export_data, ensure_ascii=False, indent=2)
        zipf.writestr("api_catalog.json", json_content)
        # Ghi ảnh
        added = set()
        for path_in_zip, path_on_disk in image_files:
            if path_in_zip not in added:
                zipf.write(path_on_disk, path_in_zip)
                added.add(path_in_zip)

    return send_file(zip_path, as_attachment=True, download_name="api_catalog_backup.zip")


@app.route("/api/image-api/import", methods=["POST"])
def api_import_api_entries():
    """Import API catalog từ file ZIP (JSON + ảnh) hoặc file JSON thuần."""
    import json
    import zipfile as zf

    if "file" not in request.files:
        return jsonify({"success": False, "error": "Chưa chọn file"}), 400

    file = request.files["file"]
    filename = file.filename.lower()

    if filename.endswith(".zip"):
        # Import từ ZIP (có ảnh)
        import tempfile
        tmp_path = os.path.join(UPLOAD_FOLDER, "import_api.zip")
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(tmp_path)

        try:
            with zf.ZipFile(tmp_path, 'r') as zipf:
                # Đọc JSON
                json_content = zipf.read("api_catalog.json").decode("utf-8")
                records = json.loads(json_content)

                # Giải nén ảnh vào thư mục static/uploads/api_images
                imported = 0
                skipped = 0
                for r in records:
                    name = r.get('name', '').strip()
                    endpoint = r.get('endpoint', '').strip()
                    method = r.get('method', 'GET').strip().upper()
                    if not name or not endpoint:
                        skipped += 1
                        continue

                    # Kiểm tra trùng
                    from database import get_db
                    conn = get_db()
                    cursor = conn.cursor()
                    cursor.execute('SELECT COUNT(*) FROM api_catalog WHERE endpoint = ? AND method = ?', (endpoint, method))
                    if cursor.fetchone()[0] > 0:
                        conn.close()
                        skipped += 1
                        continue
                    conn.close()

                    # Extract ảnh chính
                    image_path = ""
                    if r.get('image_path') and r['image_path'] in zipf.namelist():
                        import uuid
                        ext = r['image_path'].rsplit('.', 1)[1].lower() if '.' in r['image_path'] else 'png'
                        fname = f"{uuid.uuid4().hex}.{ext}"
                        dest = os.path.join(API_IMAGES_FOLDER, fname)
                        with zipf.open(r['image_path']) as src, open(dest, 'wb') as dst:
                            dst.write(src.read())
                        image_path = f"uploads/api_images/{fname}"

                    # Thêm entry
                    new_id = add_api_entry(
                        name, method, endpoint,
                        r.get('description', ''),
                        r.get('request_body', ''),
                        r.get('response_body', ''),
                        image_path,
                        r.get('category', ''),
                        r.get('notes', ''),
                        r.get('sort_order', 0)
                    )

                    # Extract ảnh phụ
                    for extra_path in r.get('extra_images', []):
                        if extra_path in zipf.namelist():
                            import uuid
                            ext = extra_path.rsplit('.', 1)[1].lower() if '.' in extra_path else 'png'
                            fname = f"{uuid.uuid4().hex}.{ext}"
                            dest = os.path.join(API_IMAGES_FOLDER, fname)
                            with zipf.open(extra_path) as src, open(dest, 'wb') as dst:
                                dst.write(src.read())
                            add_api_image(new_id, f"uploads/api_images/{fname}")

                    imported += 1

            # Cleanup
            os.remove(tmp_path)
            return jsonify({"success": True, "message": f"Đã import {imported} API (kèm ảnh), bỏ qua {skipped} trùng"})

        except Exception as e:
            return jsonify({"success": False, "error": f"Lỗi đọc file ZIP: {str(e)}"}), 400

    elif filename.endswith(".json"):
        # Import từ JSON thuần (không có ảnh)
        try:
            content = file.read().decode("utf-8")
            records = json.loads(content)
            imported, skipped = import_api_entries(records)
            return jsonify({"success": True, "message": f"Đã import {imported} API, bỏ qua {skipped} trùng"})
        except Exception as e:
            return jsonify({"success": False, "error": f"Lỗi đọc file: {str(e)}"}), 400

    else:
        return jsonify({"success": False, "error": "Chỉ hỗ trợ file .zip hoặc .json"}), 400


# ============================================================
# CHẠY APP
# Local: py app.py → http://localhost:5000
# Deploy (Render): tự dùng gunicorn
# ============================================================
if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "true").lower() == "true"
    app.run(debug=debug, host="0.0.0.0", port=port)